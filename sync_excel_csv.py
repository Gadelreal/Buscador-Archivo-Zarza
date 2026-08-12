#!/usr/bin/env python3
"""
Script de sincronización bidireccional entre CSV y Excel (.xlsx)
para el catálogo Nivel 9 (Documento simple) - Fondo Rafael Zarza
"""

import sys
import os
import csv
import argparse

def csv_to_excel(csv_path="Nivel 9 (Documento simple).csv", xlsx_path="Nivel 9 (Documento simple).xlsx"):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not os.path.exists(csv_path):
        print(f"[ERROR] No se encuentra el archivo fuente CSV: {csv_path}")
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Nivel 9 (Documento simple)"

    # Leer CSV
    with open(csv_path, mode='r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    # Estilos para cabecera
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    # Formatear cabecera (Fila 1)
    ws.row_dimensions[1].height = 28
    for col_idx, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Inmovilizar la fila superior y añadir filtros
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Formatear celdas de datos y ajustar anchos de columna
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 22
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = cell_font
            cell.border = thin_border
            # Ajustar alineación según tipo de columna
            if col_idx in [1, 2, 3, 6, 7, 8, 9]:  # Campos cortos/identificadores
                cell.alignment = Alignment(vertical="center", horizontal="left")
            else:  # Título, descripción, formato, materia
                cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    # Anchos predefinidos por columna para una visualización óptima
    col_widths = {
        1: 42, # source
        2: 44, # identifier
        3: 15, # isPartOf
        4: 45, # title
        5: 65, # description
        6: 18, # date
        7: 18, # relation
        8: 20, # coverage
        9: 16, # type
        10: 35, # format
        11: 30, # contributor
        12: 45, # subject
        13: 40  # Notas
    }

    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    wb.save(xlsx_path)
    print(f"[ÉXITO] Archivo Excel generado correctamente: {xlsx_path}")
    return True


def excel_to_csv(xlsx_path="Nivel 9 (Documento simple).xlsx", csv_path="Nivel 9 (Documento simple).csv"):
    import openpyxl

    if not os.path.exists(xlsx_path):
        print(f"[ERROR] No se encuentra el archivo fuente Excel: {xlsx_path}")
        return False

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    with open(csv_path, mode='w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f, lineterminator='\n')
        for row in ws.iter_rows(values_only=True):
            # Limpiar valores None a string vacío
            cleaned_row = [str(val) if val is not None else "" for val in row]
            # Omitir filas completamente vacías al final
            if any(cleaned_row):
                writer.writerow(cleaned_row)

    print(f"[ÉXITO] Archivo CSV sincronizado correctamente desde Excel: {csv_path}")
    return True


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Sincronizador CSV <-> Excel para Archivo Rafael Zarza")
    parser.add_argument("action", choices=["to-excel", "to-csv"], help="Acción a realizar: to-excel (CSV -> XLSX) o to-csv (XLSX -> CSV)")
    args = parser.parse_args()

    if args.action == "to-excel":
        csv_to_excel()
    elif args.action == "to-csv":
        excel_to_csv()
