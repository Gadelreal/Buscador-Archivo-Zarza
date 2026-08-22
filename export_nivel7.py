import openpyxl
import csv
import re
import sys
import os

def export_nivel7():
    if os.path.exists("Nivel 7.xlsx"):
        excel_path = "Nivel 7.xlsx"
    else:
        excel_path = "Descripcion Archivo/Nivel 7.xlsx"
    csv_path = "Nivel 7 (Carpetilla simple).csv"

    print(f"Leyendo {excel_path}...")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    if "Nivel 7" not in wb.sheetnames:
        print("Error: No se encuentra la hoja Nivel 7")
        sys.exit(1)
        
    ws = wb["Nivel 7"]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows:
        print("Error: La hoja Nivel 7 está vacía")
        sys.exit(1)
        
    # Output headers for Nivel 7 CSV
    header = [
        "source_code",      # Código Local (ej. PUBLICACIONES-C001-P001)
        "identifier",       # 1.1 Código ISAD(G) (ej. ES-CEDCS-ZARZA-PUBLICACIONES-C001-P001)
        "rotulo",           # Rótulo Físico (ej. C001, C001-P001)
        "title",            # 1.2 Título Atribuido
        "date",             # 1.3 Fecha(s)
        "format",           # 1.5 Extensión y Soporte
        "description",      # 3.1 Alcance y Contenido
        "relation"          # Serie deducida (PUBLICACIONES, EXPOSICIONES, MILITANCIA, etc.)
    ]
    
    out_rows = [header]
    
    # Process data rows (skipping header row 0)
    for i, r in enumerate(rows[1:], start=2):
        if not r or not any(r):
            continue
            
        cod_local = str(r[0]).strip() if r[0] is not None else ""
        isad_code = str(r[1]).strip() if r[1] is not None else ""
        rotulo = str(r[2]).strip() if r[2] is not None else ""
        title = str(r[3]).strip() if r[3] is not None else ""

        if not rotulo or rotulo == "#VALUE!" or not title or title == "#VALUE!":
            continue
        date = str(r[4]).strip() if r[4] is not None else ""
        ext_soporte = str(r[5]).strip() if r[5] is not None else ""
        alcance = str(r[6]).strip() if len(r) > 6 and r[6] is not None else ""
        
        # Deduce series from isad_code or cod_local
        series = "XXXXX"
        for s in ["PUBLICACIONES", "EXPOSICIONES", "MILITANCIA", "PERSONAL", "AUDIOVISUAL", "ARQUITECTURA"]:
            if s in isad_code or s in cod_local:
                series = s
                break
                
        out_rows.append([
            cod_local,
            isad_code,
            rotulo,
            title,
            date,
            ext_soporte,
            alcance,
            series
        ])
        
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerows(out_rows)
        
    print(f"[ÉXITO] Exportados {len(out_rows)-1} registros a {csv_path}")

if __name__ == "__main__":
    export_nivel7()
